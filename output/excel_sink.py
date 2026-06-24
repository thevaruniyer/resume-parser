"""
Phase 4: ExcelSink — openpyxl-based upsert into the client workbook.

Two sheets:
  "Candidates" — clean records (needs_review=False, not dead-lettered)
  "Review"     — low-confidence / flagged / dead-lettered records

Upsert semantics: match by dedup_key → update row in place; else append.
The workbook is opened and saved on every write (no persistent handle), which
prevents corruption and makes the file always readable between runs.
"""
from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from output.base import Sink
from output.column_map import COLUMN_HEADERS, COLUMN_MAP, record_to_row

logger = logging.getLogger(__name__)

_DEAD_LETTER_PATHS = {"dead_letter", "DEAD_LETTER", "dead-letter"}


class ExcelSink(Sink):
    """
    Upsert-capable Excel sink.

    Records with needs_review=True or path_taken in _DEAD_LETTER_PATHS
    are written to the Review sheet; all others go to the Candidates sheet.
    """

    MAIN_SHEET = "Candidates"
    REVIEW_SHEET = "Review"

    def __init__(self, filepath: str | Path) -> None:
        self._path = Path(filepath)
        self._path.parent.mkdir(parents=True, exist_ok=True)

        if not self._path.exists():
            self._create_workbook()
        else:
            self._ensure_sheets()

    # ------------------------------------------------------------------
    # Workbook initialisation
    # ------------------------------------------------------------------

    def _create_workbook(self) -> None:
        wb = Workbook()
        ws_main = wb.active
        ws_main.title = self.MAIN_SHEET
        _write_headers(ws_main)
        ws_review = wb.create_sheet(title=self.REVIEW_SHEET)
        _write_headers(ws_review)
        wb.save(self._path)

    def _ensure_sheets(self) -> None:
        wb = load_workbook(self._path)
        changed = False
        if self.MAIN_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(title=self.MAIN_SHEET, index=0)
            _write_headers(ws)
            changed = True
        if self.REVIEW_SHEET not in wb.sheetnames:
            ws = wb.create_sheet(title=self.REVIEW_SHEET)
            _write_headers(ws)
            changed = True
        if changed:
            wb.save(self._path)

    # ------------------------------------------------------------------
    # Core write
    # ------------------------------------------------------------------

    def write(self, record: Any) -> None:
        if hasattr(record, "model_dump"):
            record = record.model_dump()

        meta = record.get("meta") or {}
        dedup_key: Optional[str] = meta.get("dedup_key")
        needs_review: bool = bool(meta.get("needs_review", False))
        path_taken: str = str(meta.get("path_taken") or "")

        is_review = needs_review or path_taken in _DEAD_LETTER_PATHS
        sheet_name = self.REVIEW_SHEET if is_review else self.MAIN_SHEET

        row_values = record_to_row(record)

        wb = load_workbook(self._path)
        ws = wb[sheet_name]

        existing_row = _find_row_by_dedup_key(ws, dedup_key) if dedup_key else None

        if existing_row is not None:
            for col_idx, value in enumerate(row_values, 1):
                ws.cell(row=existing_row, column=col_idx, value=value)
            logger.debug("Upserted row %d in %s for dedup_key=%s", existing_row, sheet_name, dedup_key)
        else:
            ws.append(row_values)
            logger.debug("Appended to %s (dedup_key=%s)", sheet_name, dedup_key)

        wb.save(self._path)

    def close(self) -> None:
        pass  # no persistent handle to release

    # ------------------------------------------------------------------
    # Helpers used by run_batch
    # ------------------------------------------------------------------

    def get_processed_source_files(self) -> set[str]:
        """Return all source_file values already present in either sheet."""
        if not self._path.exists():
            return set()
        src_header = COLUMN_MAP["source_file"]
        processed: set[str] = set()
        wb = load_workbook(self._path, read_only=True)
        for sname in (self.MAIN_SHEET, self.REVIEW_SHEET):
            if sname not in wb.sheetnames:
                continue
            ws = wb[sname]
            src_col: Optional[int] = None
            for cell in ws[1]:
                if cell.value == src_header:
                    src_col = cell.column
                    break
            if src_col is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[src_col - 1]
                if val:
                    processed.add(str(val))
        wb.close()
        return processed

    def row_count(self, sheet: Optional[str] = None) -> int:
        """Data row count (header excluded) for one sheet or both combined."""
        if not self._path.exists():
            return 0
        names = [sheet] if sheet else [self.MAIN_SHEET, self.REVIEW_SHEET]
        total = 0
        wb = load_workbook(self._path, read_only=True)
        for sname in names:
            if sname in wb.sheetnames:
                ws = wb[sname]
                total += max(0, (ws.max_row or 1) - 1)
        wb.close()
        return total

    def get_cell_value(self, sheet: str, data_row: int, header: str) -> Any:
        """Read a cell by sheet, 1-based data row, and column header."""
        wb = load_workbook(self._path, read_only=True)
        ws = wb[sheet]
        col_idx: Optional[int] = None
        for cell in ws[1]:
            if cell.value == header:
                col_idx = cell.column
                break
        value = ws.cell(row=data_row + 1, column=col_idx).value if col_idx else None
        wb.close()
        return value


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------

def _write_headers(ws) -> None:
    ws.append(COLUMN_HEADERS)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
    ws.freeze_panes = "A2"


def _find_row_by_dedup_key(ws, dedup_key: str) -> Optional[int]:
    """Return the row number of an existing record with this dedup_key, or None."""
    target_header = COLUMN_MAP["dedup_key"]
    col_idx: Optional[int] = None
    for cell in ws[1]:
        if cell.value == target_header:
            col_idx = cell.column
            break
    if col_idx is None:
        return None
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[col_idx - 1]
        if cell.value is not None and str(cell.value) == str(dedup_key):
            return cell.row
    return None
