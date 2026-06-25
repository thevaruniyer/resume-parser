"""
Phase 5 tests — StorageConnector interface, LocalFolderConnector,
RcloneConnector (subprocess mocked), orchestration, dead-letter, idempotency.

Zero Gemini API calls.  RcloneConnector tests mock subprocess.run.
"""
from __future__ import annotations

import hashlib
import json
import os
import sys
import textwrap
from datetime import datetime, timezone
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from connectors.base import ConfigError, FileRecord, StorageConnector
from connectors.local_connector import LocalFolderConnector
from connectors.rclone_connector import RcloneConnector
from output.column_map import COLUMN_HEADERS, COLUMN_MAP, _EXTRACTORS, _most_recent_job


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def tmp_folder(tmp_path) -> Path:
    """A folder with three fake resume files."""
    for name, content in [
        ("alice.pdf", b"%PDF-1.4 fake pdf content alice"),
        ("bob.txt", b"Bob Smith resume"),
        ("carol.png", b"\x89PNG fake image carol"),
    ]:
        (tmp_path / name).write_bytes(content)
    return tmp_path


@pytest.fixture
def wb_path(tmp_path) -> Path:
    return tmp_path / "test5.xlsx"


def _make_file_record(name: str = "resume.pdf", path: str = "/tmp/resume.pdf",
                      file_hash: str = "abc123") -> FileRecord:
    return FileRecord(
        name=name,
        path=path,
        file_type="pdf",
        size=1000,
        modified_at="2026-01-01T00:00:00Z",
        file_hash=file_hash,
    )


def _parsed_record(email: str = "test@example.com", name: str = "Test User",
                   source_file: str = "resume.pdf") -> dict:
    dedup = hashlib.sha1(email.lower().encode()).hexdigest()
    return {
        "full_name": name,
        "first_name": name.split()[0],
        "last_name": name.split()[-1],
        "emails": [email],
        "phones": ["+91-98765-43210"],
        "work_experience": [{
            "company": "TestCorp",
            "designation": "Engineer",
            "start_date": "2022-01",
            "end_date": None,
            "is_current": True,
            "duration_months": 24,
        }],
        "education": [],
        "qualifications": [],
        "articleship_internships": [],
        "skills": [{"name": "Python", "category": "technical"}],
        "projects": [],
        "achievements_awards": [],
        "publications": [],
        "extracurriculars": [],
        "languages_known": [],
        "derived": {"total_experience_years": 2.0, "current_employer": "TestCorp",
                    "current_designation": "Engineer"},
        "meta": {
            "source_file": source_file,
            "source_path": f"/some/path/{source_file}",
            "file_type": "pdf",
            "parse_timestamp": "2026-01-01T00:00:00Z",
            "model_used": "gemini-2.5-flash",
            "path_taken": "text",
            "needs_review": False,
            "review_reasons": [],
            "dedup_key": dedup,
        },
    }


# ---------------------------------------------------------------------------
# Task C — Filename as first column
# ---------------------------------------------------------------------------

class TestFilenameColumn:
    def test_filename_is_first_column_header(self, wb_path):
        from output.excel_sink import ExcelSink
        from openpyxl import load_workbook
        ExcelSink(wb_path)
        wb = load_workbook(wb_path, read_only=True)
        first_header = wb["Candidates"][1][0].value
        wb.close()
        assert first_header == "Filename", f"Expected 'Filename', got {first_header!r}"

    def test_filename_value_is_bare_name(self, wb_path):
        from output.excel_sink import ExcelSink
        from openpyxl import load_workbook
        sink = ExcelSink(wb_path)
        record = _parsed_record(source_file="john_smith.pdf")
        sink.write(record)
        wb = load_workbook(wb_path, read_only=True)
        first_col_value = wb["Candidates"].cell(2, 1).value
        wb.close()
        assert first_col_value == "john_smith.pdf"

    def test_filename_in_column_map_at_index_0(self):
        keys = list(COLUMN_MAP.keys())
        assert keys[0] == "filename"
        assert COLUMN_MAP["filename"] == "Filename"

    def test_filename_extractor_returns_source_file(self):
        record = _parsed_record(source_file="jane_doe_cv.pdf")
        val = _EXTRACTORS["filename"](record)
        assert val == "jane_doe_cv.pdf"


# ---------------------------------------------------------------------------
# Task D Bug 1 — escalation ValueError fix
# ---------------------------------------------------------------------------

class TestEscalationNoImagesForVision:
    def test_txt_no_images_dead_lettered_not_valueerror(self, tmp_path):
        """
        TEXT fails, HYBRID fails, VISION finds no images → ExtractionError
        (not ValueError), escalation returns (None, decision).
        """
        from extraction.base import ExtractionError
        from routing import FileRouter, extract_with_escalation
        from routing.base import ExtractionPath
        from schema import ResumeExtractPayload

        txt_file = tmp_path / "test.txt"
        txt_file.write_text("John Smith, Engineer", encoding="utf-8")

        class AlwaysFailExtractor:
            model_name = "mock"
            def extract(self, **_):
                raise ExtractionError("mock TEXT/HYBRID failure")

        router = FileRouter()
        extractor = AlwaysFailExtractor()

        # Must not raise — should return (None, decision)
        raw, decision = extract_with_escalation(txt_file, router, extractor, ResumeExtractPayload)
        assert raw is None, "Expected None (dead-letter), got a result"
        # Decision should record no_images_for_vision in escalation history
        history = [str(h) for h in decision.escalation_history]
        # All tiers were attempted and failed
        assert len(decision.escalation_history) >= 1

    def test_no_valueerror_propagated(self, tmp_path):
        """ValueError must NOT escape extract_with_escalation."""
        from extraction.base import ExtractionError
        from routing import FileRouter, extract_with_escalation
        from schema import ResumeExtractPayload

        txt_file = tmp_path / "cv.txt"
        txt_file.write_text("Jane Doe, Accountant", encoding="utf-8")

        class FailExtractor:
            model_name = "mock"
            def extract(self, **_):
                raise ExtractionError("mock failure")

        router = FileRouter()
        try:
            result = extract_with_escalation(txt_file, router, FailExtractor(), ResumeExtractPayload)
        except ValueError as exc:
            pytest.fail(f"ValueError escaped escalation loop: {exc}")


# ---------------------------------------------------------------------------
# Task D Bug 2 — most-recent job sort order
# ---------------------------------------------------------------------------

class TestMostRecentJobSort:
    def _record_with_jobs(self, jobs: list[dict]) -> dict:
        return {"work_experience": jobs}

    def test_is_current_true_wins_over_first_entry(self):
        record = self._record_with_jobs([
            {"company": "Old Corp", "designation": "Junior", "start_date": "2015-01",
             "end_date": "2019-12", "is_current": False},
            {"company": "Current Corp", "designation": "Senior", "start_date": "2020-01",
             "end_date": None, "is_current": True},
        ])
        job = _most_recent_job(record)
        assert job["company"] == "Current Corp"
        assert _EXTRACTORS["job1_company"](record) == "Current Corp"
        assert _EXTRACTORS["job1_title"](record) == "Senior"

    def test_is_current_true_wins_even_if_lower_start_date(self):
        """Current job listed second with earlier start_date must still win."""
        record = self._record_with_jobs([
            {"company": "Higher Start", "designation": "Analyst", "start_date": "2023-01",
             "end_date": "2024-06", "is_current": False},
            {"company": "Current But Earlier", "designation": "Manager", "start_date": "2019-01",
             "end_date": None, "is_current": True},
        ])
        assert _EXTRACTORS["job1_company"](record) == "Current But Earlier"

    def test_no_current_falls_back_to_highest_start_date(self):
        record = self._record_with_jobs([
            {"company": "Earlier", "designation": "Analyst", "start_date": "2018-01",
             "end_date": "2020-12", "is_current": False},
            {"company": "Latest", "designation": "Senior", "start_date": "2022-06",
             "end_date": "2024-01", "is_current": False},
        ])
        assert _EXTRACTORS["job1_company"](record) == "Latest"

    def test_empty_work_experience_returns_none(self):
        record = {"work_experience": []}
        assert _EXTRACTORS["job1_company"](record) is None
        assert _most_recent_job(record) == {}


# ---------------------------------------------------------------------------
# LocalFolderConnector
# ---------------------------------------------------------------------------

class TestLocalFolderConnector:
    def test_list_files_returns_file_records(self, tmp_folder):
        conn = LocalFolderConnector(tmp_folder)
        records = conn.list_files()
        assert len(records) == 3
        names = {r.name for r in records}
        assert names == {"alice.pdf", "bob.txt", "carol.png"}

    def test_list_files_hash_non_null(self, tmp_folder):
        conn = LocalFolderConnector(tmp_folder)
        for fr in conn.list_files():
            assert fr.file_hash, f"{fr.name} has empty hash"
            assert len(fr.file_hash) == 64  # SHA-256 hex

    def test_list_files_file_type_correct(self, tmp_folder):
        conn = LocalFolderConnector(tmp_folder)
        types = {r.name: r.file_type for r in conn.list_files()}
        assert types["alice.pdf"] == "pdf"
        assert types["bob.txt"] == "txt"
        assert types["carol.png"] == "png"

    def test_download_returns_original_path(self, tmp_folder):
        conn = LocalFolderConnector(tmp_folder)
        fr = conn.list_files()[0]
        returned = conn.download(fr)
        assert returned == Path(fr.path)

    def test_delta_empty_manifest_returns_all(self, tmp_folder):
        conn = LocalFolderConnector(tmp_folder)
        delta = conn.delta({})
        assert len(delta) == 3

    def test_delta_full_matching_manifest_returns_empty(self, tmp_folder):
        conn = LocalFolderConnector(tmp_folder)
        records = conn.list_files()
        manifest = {r.name: {"hash": r.file_hash, "processed_at": "now"} for r in records}
        delta = conn.delta(manifest)
        assert delta == []

    def test_delta_one_hash_changed_returns_only_that_file(self, tmp_folder):
        conn = LocalFolderConnector(tmp_folder)
        records = conn.list_files()
        manifest = {r.name: {"hash": r.file_hash, "processed_at": "now"} for r in records}
        # Corrupt alice's hash in manifest
        manifest["alice.pdf"]["hash"] = "stale_hash"
        delta = conn.delta(manifest)
        assert len(delta) == 1
        assert delta[0].name == "alice.pdf"

    def test_save_and_load_manifest(self, tmp_folder, tmp_path):
        conn = LocalFolderConnector(tmp_folder)
        # Override manifest path to tmp
        conn._manifest_path = tmp_path / "test_manifest.json"
        data = {"resume.pdf": {"hash": "abc", "processed_at": "2026-01-01T00:00:00Z"}}
        conn.save_manifest(data)
        loaded = conn.load_manifest()
        assert loaded == data

    def test_load_manifest_returns_empty_when_missing(self, tmp_folder, tmp_path):
        conn = LocalFolderConnector(tmp_folder)
        conn._manifest_path = tmp_path / "nonexistent.json"
        assert conn.load_manifest() == {}

    def test_cleanup_is_noop(self, tmp_folder):
        """LocalFolderConnector.cleanup_downloaded must not delete the file."""
        conn = LocalFolderConnector(tmp_folder)
        target = tmp_folder / "alice.pdf"
        conn.cleanup_downloaded(target)
        assert target.exists()


# ---------------------------------------------------------------------------
# RcloneConnector (subprocess mocked)
# ---------------------------------------------------------------------------

MOCK_LSJSON = json.dumps([
    {"Name": "resume1.pdf",  "Path": "ResumeTest/resume1.pdf",  "Size": 12345,
     "ModTime": "2026-01-01T00:00:00Z", "Hashes": {"SHA-1": "sha1abc"}, "IsDir": False},
    {"Name": "resume2.jpg",  "Path": "ResumeTest/resume2.jpg",  "Size": 67890,
     "ModTime": "2026-01-02T00:00:00Z", "Hashes": {"SHA-1": "sha1def"}, "IsDir": False},
    {"Name": "cv_md5only.png","Path": "ResumeTest/cv_md5only.png","Size": 5000,
     "ModTime": "2026-01-03T00:00:00Z", "Hashes": {"MD5": "md5ghi"}, "IsDir": False},
    {"Name": "subfolder",     "Path": "ResumeTest/subfolder",    "Size": 0,
     "ModTime": "2026-01-01T00:00:00Z", "Hashes": {},              "IsDir": True},
])


def _mock_run_lsjson(cmd, **_):
    m = MagicMock()
    m.stdout = MOCK_LSJSON
    m.returncode = 0
    return m


def _mock_run_copyto(cmd, **_):
    # Create the destination file so the path exists
    dest = Path(cmd[-1])
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(b"fake downloaded content")
    m = MagicMock()
    m.stdout = ""
    m.returncode = 0
    return m


class TestRcloneConnector:
    def test_list_files_filters_dirs(self):
        with patch("connectors.rclone_connector._run", side_effect=_mock_run_lsjson):
            conn = RcloneConnector(remote="onedrive", path="ResumeTest/")
            records = conn.list_files()
        assert len(records) == 3
        names = {r.name for r in records}
        assert "subfolder" not in names

    def test_list_files_sha1_hash_used(self):
        with patch("connectors.rclone_connector._run", side_effect=_mock_run_lsjson):
            conn = RcloneConnector(remote="onedrive", path="ResumeTest/")
            records = conn.list_files()
        by_name = {r.name: r for r in records}
        assert by_name["resume1.pdf"].file_hash == "sha1abc"
        assert by_name["resume2.jpg"].file_hash == "sha1def"

    def test_list_files_falls_back_to_md5(self):
        with patch("connectors.rclone_connector._run", side_effect=_mock_run_lsjson):
            conn = RcloneConnector(remote="onedrive", path="ResumeTest/")
            records = conn.list_files()
        by_name = {r.name: r for r in records}
        assert by_name["cv_md5only.png"].file_hash == "md5ghi"

    def test_download_calls_correct_rclone_command(self, tmp_path):
        captured_cmds = []

        def mock_run(cmd, **_):
            captured_cmds.append(cmd)
            dest = Path(cmd[-1])
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(b"fake")
            m = MagicMock()
            m.returncode = 0
            m.stdout = ""
            return m

        conn = RcloneConnector(remote="onedrive", path="ResumeTest/")
        conn._manifest_path = tmp_path / "manifest.json"
        fr = FileRecord(name="resume1.pdf", path="ResumeTest/resume1.pdf",
                        file_type="pdf", file_hash="abc")

        with patch("connectors.rclone_connector._run", side_effect=mock_run):
            local = conn.download(fr)

        assert len(captured_cmds) == 1
        cmd = captured_cmds[0]
        assert cmd[0] == "rclone"
        assert cmd[1] == "copyto"
        assert "onedrive:ResumeTest/resume1.pdf" in cmd
        assert local.name == "resume1.pdf"

    def test_cleanup_deletes_tmp_file(self, tmp_path):
        conn = RcloneConnector(remote="onedrive", path="ResumeTest/")
        fake = tmp_path / "fake.pdf"
        fake.write_bytes(b"content")
        assert fake.exists()
        conn.cleanup_downloaded(fake)
        assert not fake.exists()

    def test_missing_remote_raises_config_error(self, monkeypatch):
        monkeypatch.setenv("RCLONE_REMOTE", "")
        with pytest.raises(ConfigError, match="RCLONE_REMOTE"):
            RcloneConnector(remote="")

    def test_delta_uses_file_hash(self):
        with patch("connectors.rclone_connector._run", side_effect=_mock_run_lsjson):
            conn = RcloneConnector(remote="onedrive", path="ResumeTest/")
            manifest = {"resume1.pdf": {"hash": "sha1abc", "processed_at": "now"}}
            delta = conn.delta(manifest)
        # resume1.pdf is unchanged; resume2.jpg and cv_md5only.png are new
        names = {fr.name for fr in delta}
        assert "resume1.pdf" not in names
        assert "resume2.jpg" in names
        assert "cv_md5only.png" in names


# ---------------------------------------------------------------------------
# Orchestration tests (mock connector + mock extractor)
# ---------------------------------------------------------------------------

def _make_mock_connector(tmp_path: Path, records: list[FileRecord]) -> MagicMock:
    conn = MagicMock(spec=StorageConnector)
    conn.load_manifest.return_value = {}
    conn.list_files.return_value = records
    conn.delta.return_value = records
    # download: write a fake file and return its path
    def fake_download(fr):
        p = tmp_path / fr.name
        p.write_text("fake content for " + fr.name, encoding="utf-8")
        return p
    conn.download.side_effect = fake_download
    conn.cleanup_downloaded = MagicMock()
    conn.save_manifest = MagicMock()
    return conn


def _make_mock_extractor(record_override: dict | None = None) -> MagicMock:
    from schema import ResumeExtractPayload
    ext = MagicMock()
    ext.model_name = "mock-model"
    payload = {
        "full_name": "Rahul Mehta",
        "first_name": "Rahul",
        "last_name": "Mehta",
        "emails": ["rahul.mehta@example.com"],
        "phones": ["+91-98765-43210"],
        "work_experience": [],
        "education": [],
        "qualifications": [],
        "articleship_internships": [],
        "skills": [],
        "projects": [],
        "achievements_awards": [],
        "publications": [],
        "extracurriculars": [],
        "languages_known": [],
    }
    if record_override:
        payload.update(record_override)
    ext.extract.return_value = payload
    return ext


def _mock_gemini_extractor():
    """Return a MagicMock that looks enough like GeminiExtractor for run_batch."""
    ext = MagicMock()
    ext.model_name = "mock-model"
    return ext


def _mock_extract_with_escalation(local_path, router, extractor, schema, **kwargs):
    """Return a parsed dict without touching the file (for orchestration tests)."""
    from routing.base import ExtractionPath, RoutingDecision
    decision = RoutingDecision(
        path=ExtractionPath.TEXT,
        file_type=local_path.suffix.lstrip("."),
        reason="mock",
    )
    payload = {
        "full_name": "Rahul Mehta",
        "first_name": "Rahul",
        "last_name": "Mehta",
        "emails": ["rahul.mehta@example.com"],
        "phones": ["+91-98765-43210"],
        "work_experience": [],
        "education": [],
        "qualifications": [],
        "articleship_internships": [],
        "skills": [],
        "projects": [],
        "achievements_awards": [],
        "publications": [],
        "extracurriculars": [],
        "languages_known": [],
    }
    return payload, decision


class TestOrchestration:
    def test_writes_rows_to_excel(self, tmp_path, wb_path):
        from output.excel_sink import ExcelSink
        from run_batch import run_batch

        fr = _make_file_record("resume.pdf", str(tmp_path / "resume.pdf"))
        conn = _make_mock_connector(tmp_path, [fr])

        with patch("run_batch._build_connector", return_value=conn), \
             patch("run_batch.build_fallback_extractor", return_value=_mock_gemini_extractor()), \
             patch("run_batch.extract_with_escalation", side_effect=_mock_extract_with_escalation):
            summary = run_batch(output_path=wb_path, connector_type="local", verbose=False)

        assert summary["parsed_ok"] == 1
        sink = ExcelSink(wb_path)
        assert sink.row_count("Candidates") >= 1

    def test_filename_is_first_column_in_written_row(self, tmp_path, wb_path):
        from openpyxl import load_workbook
        from run_batch import run_batch

        fr = _make_file_record("john_smith.pdf", str(tmp_path / "john_smith.pdf"))
        conn = _make_mock_connector(tmp_path, [fr])

        with patch("run_batch._build_connector", return_value=conn), \
             patch("run_batch.build_fallback_extractor", return_value=_mock_gemini_extractor()), \
             patch("run_batch.extract_with_escalation", side_effect=_mock_extract_with_escalation):
            run_batch(output_path=wb_path, connector_type="local", verbose=False)

        wb = load_workbook(wb_path, read_only=True)
        first_col_value = wb["Candidates"].cell(2, 1).value
        wb.close()
        assert first_col_value == "john_smith.pdf"

    def test_manifest_updated_after_run(self, tmp_path, wb_path):
        from run_batch import run_batch

        fr = _make_file_record("resume.pdf", str(tmp_path / "resume.pdf"), file_hash="hash123")
        conn = _make_mock_connector(tmp_path, [fr])

        with patch("run_batch._build_connector", return_value=conn), \
             patch("run_batch.build_fallback_extractor", return_value=_mock_gemini_extractor()), \
             patch("run_batch.extract_with_escalation", side_effect=_mock_extract_with_escalation):
            run_batch(output_path=wb_path, connector_type="local", verbose=False)

        conn.save_manifest.assert_called()
        saved_manifest = conn.save_manifest.call_args[0][0]
        assert "resume.pdf" in saved_manifest
        assert saved_manifest["resume.pdf"]["hash"] == "hash123"

    def test_tmp_cleanup_called(self, tmp_path, wb_path):
        from run_batch import run_batch

        fr = _make_file_record("resume.pdf", str(tmp_path / "resume.pdf"))
        conn = _make_mock_connector(tmp_path, [fr])

        with patch("run_batch._build_connector", return_value=conn), \
             patch("run_batch.build_fallback_extractor", return_value=_mock_gemini_extractor()), \
             patch("run_batch.extract_with_escalation", side_effect=_mock_extract_with_escalation):
            run_batch(output_path=wb_path, connector_type="local", verbose=False)

        conn.cleanup_downloaded.assert_called_once()


# ---------------------------------------------------------------------------
# Dead-letter
# ---------------------------------------------------------------------------

class TestDeadLetter:
    def test_extractor_failure_writes_dead_letter_jsonl(self, tmp_path, wb_path):
        from run_batch import run_batch

        fr = _make_file_record("bad.pdf", str(tmp_path / "bad.pdf"))
        conn = _make_mock_connector(tmp_path, [fr])

        def always_raise(*a, **kw):
            raise RuntimeError("mock extraction failure")

        dl_log = tmp_path / "dead_letter.jsonl"

        with patch("run_batch._build_connector", return_value=conn), \
             patch("run_batch.build_fallback_extractor"), \
             patch("run_batch.extract_with_escalation", side_effect=always_raise), \
             patch("run_batch._DEAD_LETTER_LOG", dl_log):
            summary = run_batch(output_path=wb_path, connector_type="local", verbose=False)

        assert summary["dead_lettered"] >= 1
        assert dl_log.exists()
        lines = [json.loads(l) for l in dl_log.read_text().strip().splitlines()]
        assert any(entry["filename"] == "bad.pdf" for entry in lines)

    def test_dead_lettered_file_goes_to_review_sheet(self, tmp_path, wb_path):
        from output.excel_sink import ExcelSink
        from run_batch import run_batch

        fr = _make_file_record("bad.pdf", str(tmp_path / "bad.pdf"))
        conn = _make_mock_connector(tmp_path, [fr])

        def always_raise(*a, **kw):
            raise RuntimeError("mock extraction failure")

        with patch("run_batch._build_connector", return_value=conn), \
             patch("run_batch.build_fallback_extractor", return_value=_mock_gemini_extractor()), \
             patch("run_batch.extract_with_escalation", side_effect=always_raise):
            run_batch(output_path=wb_path, connector_type="local", verbose=False)

        sink = ExcelSink(wb_path)
        assert sink.row_count("Candidates") == 0
        assert sink.row_count("Review") >= 1


# ---------------------------------------------------------------------------
# Idempotency
# ---------------------------------------------------------------------------

class TestIdempotency:
    def test_second_run_with_empty_delta_adds_no_rows(self, tmp_path, wb_path):
        from output.excel_sink import ExcelSink
        from run_batch import run_batch

        fr = _make_file_record("resume.pdf", str(tmp_path / "resume.pdf"), file_hash="h1")
        conn = _make_mock_connector(tmp_path, [fr])

        with patch("run_batch._build_connector", return_value=conn), \
             patch("run_batch.build_fallback_extractor", return_value=_mock_gemini_extractor()), \
             patch("run_batch.extract_with_escalation", side_effect=_mock_extract_with_escalation):
            run_batch(output_path=wb_path, connector_type="local", verbose=False)

        row_count_after_first = ExcelSink(wb_path).row_count("Candidates")

        # Second run: manifest already has the file → delta returns nothing
        conn2 = MagicMock(spec=StorageConnector)
        conn2.load_manifest.return_value = {
            "resume.pdf": {"hash": "h1", "processed_at": "2026-01-01T00:00:00Z"}
        }
        conn2.delta.return_value = []  # nothing new

        with patch("run_batch._build_connector", return_value=conn2), \
             patch("run_batch.build_fallback_extractor", return_value=_mock_gemini_extractor()), \
             patch("run_batch.extract_with_escalation", side_effect=_mock_extract_with_escalation):
            summary2 = run_batch(output_path=wb_path, connector_type="local", verbose=False)

        assert summary2["files_in_delta"] == 0
        assert ExcelSink(wb_path).row_count("Candidates") == row_count_after_first
