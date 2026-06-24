"""
Phase 4 tests — Excel Sink + idempotent upsert.

All tests are pure Python (no Gemini API calls) except TestEndToEnd which
requires GEMINI_API_KEY and is skipped automatically when absent.
"""
from __future__ import annotations

import hashlib
import json
import os
import sys
from pathlib import Path

import pytest
from dotenv import load_dotenv

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))
load_dotenv(ROOT / ".env")

GT_FILE = ROOT / "ground_truth" / "sample_ca_india_synthetic.json"
CORPUS = ROOT / "test_corpus" / "files"

from normalize import normalize_record
from output.column_map import COLUMN_HEADERS, COLUMN_MAP, record_to_row
from output.excel_sink import ExcelSink
from validate import validate_record

_GEMINI_KEY = bool(os.getenv("GEMINI_API_KEY"))
_SKIP_API = pytest.mark.skipif(not _GEMINI_KEY, reason="GEMINI_API_KEY not set")


# ---------------------------------------------------------------------------
# Fixtures & shared helpers
# ---------------------------------------------------------------------------

@pytest.fixture
def wb_path(tmp_path) -> Path:
    return tmp_path / "test_output.xlsx"


def _load_gt() -> dict:
    """Load + normalize the CA ground-truth record (no API call needed)."""
    with open(GT_FILE) as f:
        record = json.load(f)
    record.pop("_notes", None)
    record.setdefault("meta", {})
    record["meta"].update({
        "source_file": "sample_ca_india_synthetic.txt",
        "source_path": "test_corpus/files/sample_ca_india_synthetic.txt",
        "file_type":   "txt",
        "parse_timestamp": "2024-01-01T00:00:00Z",
        "path_taken":  "text",
        "needs_review": False,
        "review_reasons": [],
    })
    return normalize_record(record)


def _non_ca_record(
    email: str = "jane.doe@example.com",
    ts: str = "2024-01-01T00:00:00Z",
) -> dict:
    dedup = hashlib.sha1(email.lower().encode()).hexdigest()
    return {
        "full_name": "Jane Doe",
        "first_name": "Jane",
        "last_name": "Doe",
        "emails": [email],
        "phones": ["+91-98765-11111"],
        "location": {"city": "Bangalore", "state": "Karnataka", "country": "India"},
        "work_experience": [{
            "company": "TechCorp India",
            "designation": "Software Engineer",
            "employment_type": "full_time",
            "start_date": "2022-01",
            "end_date": "2024-06",
            "is_current": False,
            "duration_months": 29,
        }],
        "education": [{
            "degree": "B.Tech",
            "specialization": "Computer Science",
            "institution": "IIT Delhi",
            "start_date": "2018",
            "end_date": "2022",
            "status": "completed",
            "grade_type": "CGPA",
            "grade_value": "8.5",
        }],
        "skills": [
            {"name": "Python", "category": "technical"},
            {"name": "React",  "category": "technical"},
        ],
        "qualifications": [],
        "articleship_internships": [],
        "projects": [],
        "achievements_awards": [],
        "publications": [],
        "extracurriculars": [],
        "languages_known": [{"language": "English", "proficiency": "Fluent"}],
        "derived": {
            "total_experience_years": 2.4,
            "current_employer": None,
            "current_designation": None,
        },
        "meta": {
            "source_file": "jane_doe_software_engineer.pdf",
            "source_path": "test/jane_doe_software_engineer.pdf",
            "file_type": "pdf",
            "parse_timestamp": ts,
            "model_used": "gemini-2.5-flash",
            "path_taken": "text",
            "needs_review": False,
            "review_reasons": [],
            "dedup_key": dedup,
        },
    }


def _sheet_headers(wb_path: Path, sheet_name: str) -> list[str]:
    from openpyxl import load_workbook
    wb = load_workbook(wb_path, read_only=True)
    headers = [cell.value for cell in wb[sheet_name][1]]
    wb.close()
    return headers


def _row_count(wb_path: Path, sheet_name: str) -> int:
    from openpyxl import load_workbook
    wb = load_workbook(wb_path, read_only=True)
    ws = wb[sheet_name]
    count = max(0, (ws.max_row or 1) - 1)
    wb.close()
    return count


def _cell(wb_path: Path, sheet_name: str, data_row: int, col_key: str):
    """Read a cell by logical column key; data_row is 1-based."""
    from openpyxl import load_workbook
    header = COLUMN_MAP[col_key]
    wb = load_workbook(wb_path, read_only=True)
    ws = wb[sheet_name]
    col_idx = None
    for c in ws[1]:
        if c.value == header:
            col_idx = c.column
            break
    value = ws.cell(row=data_row + 1, column=col_idx).value if col_idx else None
    wb.close()
    return value


# ---------------------------------------------------------------------------
# Column header tests
# ---------------------------------------------------------------------------

class TestColumnHeaders:
    def test_all_headers_in_main_sheet(self, wb_path):
        ExcelSink(wb_path)
        headers = _sheet_headers(wb_path, ExcelSink.MAIN_SHEET)
        for h in COLUMN_HEADERS:
            assert h in headers, f"Missing: {h!r}"

    def test_all_headers_in_review_sheet(self, wb_path):
        ExcelSink(wb_path)
        headers = _sheet_headers(wb_path, ExcelSink.REVIEW_SHEET)
        for h in COLUMN_HEADERS:
            assert h in headers, f"Missing in Review: {h!r}"

    def test_column_order_matches_column_map(self, wb_path):
        ExcelSink(wb_path)
        assert _sheet_headers(wb_path, ExcelSink.MAIN_SHEET) == COLUMN_HEADERS

    def test_column_count_matches_column_map(self, wb_path):
        ExcelSink(wb_path)
        assert len(_sheet_headers(wb_path, ExcelSink.MAIN_SHEET)) == len(COLUMN_MAP)

    def test_workbook_has_both_sheets(self, wb_path):
        from openpyxl import load_workbook
        ExcelSink(wb_path)
        wb = load_workbook(wb_path, read_only=True)
        assert ExcelSink.MAIN_SHEET in wb.sheetnames
        assert ExcelSink.REVIEW_SHEET in wb.sheetnames
        wb.close()


# ---------------------------------------------------------------------------
# Write CA record (ground truth)
# ---------------------------------------------------------------------------

class TestWriteCARecord:
    def test_goes_to_main_sheet(self, wb_path):
        sink = ExcelSink(wb_path)
        sink.write(_load_gt())
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 1
        assert _row_count(wb_path, ExcelSink.REVIEW_SHEET) == 0

    def test_full_name_correct(self, wb_path):
        ExcelSink(wb_path).write(_load_gt())
        assert _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "full_name") == "Rahul Mehta"

    def test_icai_membership_number(self, wb_path):
        ExcelSink(wb_path).write(_load_gt())
        assert _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "icai_membership_number") == "123456"

    def test_total_experience_years_positive(self, wb_path):
        ExcelSink(wb_path).write(_load_gt())
        val = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "total_experience_years")
        assert val is not None and float(val) > 0

    def test_all_skills_populated(self, wb_path):
        ExcelSink(wb_path).write(_load_gt())
        val = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "all_skills")
        assert val is not None
        # Ground truth has Tally ERP 9, SAP FICO, MS Excel, etc.
        assert any(kw in val for kw in ("Tally", "SAP", "Excel", "GST"))

    def test_ca_level_is_final(self, wb_path):
        ExcelSink(wb_path).write(_load_gt())
        assert _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "ca_level") == "Final"

    def test_dedup_key_is_sha1(self, wb_path):
        ExcelSink(wb_path).write(_load_gt())
        val = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "dedup_key")
        assert val is not None and len(str(val)) == 40

    def test_raw_json_present(self, wb_path):
        ExcelSink(wb_path).write(_load_gt())
        val = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "raw_json")
        assert val is not None
        parsed = json.loads(val)
        assert parsed.get("full_name") == "Rahul Mehta"

    def test_articleship_firm_correct(self, wb_path):
        ExcelSink(wb_path).write(_load_gt())
        val = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "articleship_firm")
        assert val == "Deloitte Haskins & Sells LLP"


# ---------------------------------------------------------------------------
# Write non-CA record
# ---------------------------------------------------------------------------

class TestWriteNonCARecord:
    def test_goes_to_main_sheet(self, wb_path):
        ExcelSink(wb_path).write(_non_ca_record())
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 1
        assert _row_count(wb_path, ExcelSink.REVIEW_SHEET) == 0

    def test_icai_membership_null(self, wb_path):
        ExcelSink(wb_path).write(_non_ca_record())
        assert _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "icai_membership_number") is None

    def test_qualification_names_null(self, wb_path):
        ExcelSink(wb_path).write(_non_ca_record())
        assert _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "qualification_names") is None

    def test_qualifications_count_zero(self, wb_path):
        ExcelSink(wb_path).write(_non_ca_record())
        assert _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "qualifications_count") == 0

    def test_job1_company_correct(self, wb_path):
        ExcelSink(wb_path).write(_non_ca_record())
        assert _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "job1_company") == "TechCorp India"

    def test_technical_skills_populated(self, wb_path):
        ExcelSink(wb_path).write(_non_ca_record())
        val = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "technical_skills")
        assert val is not None and "Python" in val

    def test_no_error_on_absent_qual_fields(self, wb_path):
        """Writing a non-CA record must not raise any exception."""
        ExcelSink(wb_path).write(_non_ca_record())  # no exception = pass


# ---------------------------------------------------------------------------
# Routing to Review sheet
# ---------------------------------------------------------------------------

class TestReviewRouting:
    def test_needs_review_true_goes_to_review(self, wb_path):
        sink = ExcelSink(wb_path)
        r = _non_ca_record(email="review@example.com")
        r["meta"]["needs_review"] = True
        r["meta"]["review_reasons"] = ["forced_for_test"]
        sink.write(r)
        assert _row_count(wb_path, ExcelSink.REVIEW_SHEET) == 1
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 0

    def test_needs_review_false_goes_to_main(self, wb_path):
        sink = ExcelSink(wb_path)
        r = _non_ca_record()
        assert r["meta"]["needs_review"] is False
        sink.write(r)
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 1
        assert _row_count(wb_path, ExcelSink.REVIEW_SHEET) == 0

    def test_dead_letter_goes_to_review(self, wb_path):
        sink = ExcelSink(wb_path)
        dl = {
            "full_name": None,
            "emails": [],
            "phones": [],
            "meta": {
                "source_file": "bad_file.doc",
                "source_path": "bad_file.doc",
                "file_type": "doc",
                "parse_timestamp": "2024-01-01T00:00:00Z",
                "path_taken": "DEAD_LETTER",
                "needs_review": True,
                "review_reasons": ["dead_letter:doc_libreoffice_not_available"],
                "dedup_key": hashlib.sha1(b"bad_file.doc").hexdigest(),
            },
        }
        sink.write(dl)
        assert _row_count(wb_path, ExcelSink.REVIEW_SHEET) == 1
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 0

    @pytest.mark.parametrize("path_val", ["DEAD_LETTER", "dead_letter"])
    def test_both_dead_letter_variants_go_to_review(self, wb_path, path_val):
        sink = ExcelSink(wb_path)
        dl = {
            "emails": [],
            "phones": [],
            "meta": {
                "source_file": f"file_{path_val}.doc",
                "source_path": ".",
                "file_type": "doc",
                "parse_timestamp": "2024-01-01T00:00:00Z",
                "path_taken": path_val,
                "needs_review": False,  # path_taken alone must trigger review
                "review_reasons": [],
                "dedup_key": hashlib.sha1(f"file_{path_val}".encode()).hexdigest(),
            },
        }
        sink.write(dl)
        assert _row_count(wb_path, ExcelSink.REVIEW_SHEET) >= 1
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 0


# ---------------------------------------------------------------------------
# Idempotency
# ---------------------------------------------------------------------------

class TestIdempotency:
    def test_same_record_row_count_unchanged(self, wb_path):
        sink = ExcelSink(wb_path)
        record = _load_gt()
        sink.write(record)
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 1
        sink.write(record)
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 1  # still 1

    def test_same_record_cell_values_unchanged(self, wb_path):
        sink = ExcelSink(wb_path)
        record = _load_gt()
        sink.write(record)
        v1 = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "full_name")
        sink.write(record)
        v2 = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "full_name")
        assert v1 == v2 == "Rahul Mehta"

    def test_different_records_no_merge(self, wb_path):
        sink = ExcelSink(wb_path)
        sink.write(_load_gt())
        sink.write(_non_ca_record())
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 2


# ---------------------------------------------------------------------------
# Upsert
# ---------------------------------------------------------------------------

class TestUpsert:
    def test_same_dedup_key_updates_not_appends(self, wb_path):
        sink = ExcelSink(wb_path)
        r1 = _non_ca_record(ts="2024-01-01T00:00:00Z")
        r2 = _non_ca_record(ts="2024-06-01T00:00:00Z")  # same email → same dedup_key
        assert r1["meta"]["dedup_key"] == r2["meta"]["dedup_key"]
        sink.write(r1)
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 1
        sink.write(r2)
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 1  # upserted, not appended

    def test_upsert_overwrites_parse_timestamp(self, wb_path):
        sink = ExcelSink(wb_path)
        sink.write(_non_ca_record(ts="2024-01-01T00:00:00Z"))
        sink.write(_non_ca_record(ts="2024-06-01T00:00:00Z"))
        ts = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "parse_timestamp")
        assert ts == "2024-06-01T00:00:00Z"

    def test_different_dedup_keys_two_rows(self, wb_path):
        sink = ExcelSink(wb_path)
        sink.write(_non_ca_record(email="alpha@example.com"))
        sink.write(_non_ca_record(email="beta@example.com"))
        assert _row_count(wb_path, ExcelSink.MAIN_SHEET) == 2


# ---------------------------------------------------------------------------
# record_to_row unit tests
# ---------------------------------------------------------------------------

class TestRecordToRow:
    def test_row_length_matches_column_map(self):
        row = record_to_row(_non_ca_record())
        assert len(row) == len(COLUMN_MAP)

    def test_dead_letter_dict_does_not_raise(self):
        dl = {
            "emails": [],
            "phones": [],
            "meta": {
                "source_file": "x.doc",
                "path_taken": "DEAD_LETTER",
                "needs_review": True,
                "review_reasons": ["dead_letter:libreoffice_missing"],
                "dedup_key": "abc123",
            },
        }
        row = record_to_row(dl)  # must not raise
        assert len(row) == len(COLUMN_MAP)

    def test_non_ca_qual_columns_are_none(self):
        row = record_to_row(_non_ca_record())
        keys = list(COLUMN_MAP.keys())
        icai_idx = keys.index("icai_membership_number")
        ca_level_idx = keys.index("ca_level")
        assert row[icai_idx] is None
        assert row[ca_level_idx] is None


# ---------------------------------------------------------------------------
# End-to-end with Gemini (API test — skipped if no key)
# ---------------------------------------------------------------------------

class TestEndToEnd:
    @_SKIP_API
    def test_pdf_and_jpg_merge_to_one_row(self, wb_path):
        """
        Process PDF + JPG of the same candidate (Rahul Mehta) through the full
        pipeline.  Both produce a record; dedup merges them to ONE main-sheet
        row; icai_membership_number == "123456".
        """
        from datetime import datetime, timezone

        from config import settings
        from extraction.gemini_extractor import GeminiExtractor
        from routing import FileRouter, extract_with_escalation
        from schema import ResumeExtractPayload, ResumeRecord

        sink = ExcelSink(wb_path)
        router = FileRouter()
        extractor = GeminiExtractor(api_key=settings.gemini_api_key)

        for fname in ("sample_ca_india_synthetic.pdf", "sample_ca_india_synthetic.jpg"):
            path = CORPUS / fname
            raw, decision = extract_with_escalation(
                path, router, extractor, ResumeExtractPayload
            )
            assert raw is not None, f"Extraction returned None for {fname}"
            raw.pop("_usage", None)

            record = ResumeRecord.model_validate({
                **raw,
                "meta": {
                    "source_file": fname,
                    "source_path": str(path),
                    "file_type": path.suffix.lstrip("."),
                    "parse_timestamp": datetime.now(timezone.utc).isoformat(),
                    "model_used": extractor.model_name,
                    "path_taken": decision.path.value,
                    "needs_review": False,
                },
            })
            rd = record.model_dump()
            rd = normalize_record(rd)
            rd = validate_record(rd)
            sink.write(rd)

        main_rows = _row_count(wb_path, ExcelSink.MAIN_SHEET)
        assert main_rows == 1, f"Expected 1 (deduped), got {main_rows}"

        icai = _cell(wb_path, ExcelSink.MAIN_SHEET, 1, "icai_membership_number")
        assert icai == "123456", f"ICAI membership expected '123456', got {icai!r}"
